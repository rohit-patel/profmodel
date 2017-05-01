#Openpyxl imports
from openpyxl import load_workbook, Workbook, cell
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook


#Database imports
from django.db import transaction
from prof.models import FileSpace, RunSpace, TransactionData, PnLData
from django.core.files.base import ContentFile

#Global constants import
from prof.settings import Transactions_settings_dict, PnL_settings_dict

#Other imports
from datetime import date, timedelta, datetime
import pandas as pd, numpy as np


def ifHeaderReturnIndex(row, desiredColumns):
    '''This function checks if a particular excel row is the header row. It does so by simply checking that each of the
    column names in the desired header row appear in the row once, and exactly once. If it is a header row, the function
    returns the index of the header columms as a tuple. If not, it retruns the python 'None' object. The arguments are:
    1. An openpyxl worksheet
    2. A tuple of strings - the column headers'''
    rowtuple=tuple(cell.value for cell in row)
    i=True
    
    for ColumnName in desiredColumns:
        if rowtuple.count(ColumnName) != 1: i=False
    if i==False:
        return None
    else:
        a=list()
        for i in desiredColumns:
            a.append(rowtuple.index(i))
        return tuple(a)
    
    
def isRowValid(row,index):
    '''Determines if the row is a valid data row. For the moment, simply checks if the first column dicted by selectedIndex is empty.'''
    return bool(row[index[0]].value)

def isRowAggregate(row,index):
    '''This function determines if a row is an aggregate row. For the moment, it will simply look at the bold rows.'''
    return row[index[0]].font.b

@transaction.atomic
def ProcessTransactionData(fileObject, run, Transactions_settings_dict = Transactions_settings_dict,foreignKeyFileObject=None):
    ''' This function takes the file object, the name of the sheet in which the data is stores, and loads the transaction data into the Django model TransactionData.'''
    sheetName = Transactions_settings_dict['TransactionSheetName']
    desiredColumns = Transactions_settings_dict['desiredColumns']
    if not(foreignKeyFileObject): foreignKeyFileObject = fileObject
    wb=load_workbook(fileObject.File, read_only=True, data_only=True)
    ws=wb.get_sheet_by_name(sheetName)
    selectedIndex=None
    for row in ws.rows:
        if selectedIndex:
            if isRowValid(row,selectedIndex):
                newrow=[row[i].value for i in selectedIndex]
                dictionary=dict(zip(outputColumnNames,newrow))
                dictionary['run']=run
                dictionary['SourceFile']=foreignKeyFileObject
                TransactionData(**dictionary).save()
        else:
            if ifHeaderReturnIndex(row,desiredColumns):
                selectedIndex=ifHeaderReturnIndex(row, desiredColumns)
                outputColumnNames=[Transactions_settings_dict['column_names_to_model_dict'][row[i].value] for i in selectedIndex]






def revCells(row):
    '''This is formatting for a row that carries revenue.'''
    for cell in row:
        cell.fill=PatternFill(start_color='d2fedb', end_color='d2fedb', fill_type='solid')

def costCells(row):
    '''This is formatting for a row that carries costs.'''
    for cell in row:
        cell.fill=PatternFill(start_color='ffe8e8', end_color='ffe8e8', fill_type='solid')

def headerCells(row):
    '''This is formatting for a row that has column names.'''
    for cell in row:
        cell.fill=PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
        cell.font=Font(bold=True)

def makeBold(row):
    '''This function makes a row bold'''
    for cell in row:
        cell.font=Font(bold=True)


def fixWidth(ws):
    '''
    This function takes an openpyxl worksheet as an input and returns an openpyxl worksheet after setting the column
    width for columns in the sheet at the maximum of the length of the string in any of the cells in the column.
    '''
    column_widths = []
    for row in ws:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value))+2 > column_widths[i]:
                    column_widths[i] = len(str(cell.value))+2
            else:
                column_widths += [len(str(cell.value))+2]

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = column_width
        
def isRevAggregate(row,index,revenueMarkers=PnL_settings_dict['revenueMarkers']):
    '''
    This function simply determined if a row in the PnL is a "Net Revenue" row, mostly used for stlying purposes.
    Presently, it simply looks at the text in the firxt column and checks if any of the markets are present.
    '''
    mark=False
    for marker in revenueMarkers:
        if marker in row[index[0]].value: mark=True
    return mark
    
def applyPnLStyling(ws,pnl_column_names,revenueMarkers):
    headerfound=False
    Revenue=True
    for row in ws.rows:
        if headerfound and Revenue:
            revCells(row)
            if isRevAggregate(row,selectedIndex,revenueMarkers): Revenue = False
        elif headerfound and not(Revenue):
            costCells(row)
        elif ifHeaderReturnIndex(row,pnl_column_names):
            headerCells(row)
            selectedIndex=ifHeaderReturnIndex(row, pnl_column_names)
            headerfound=True
        else:
            makeBold(row)
    

def generateSamplePnLBUSheet(run, ws, BU, cunstructor_dict = PnL_settings_dict):
    '''
    This function generates a sample PnL based on the transaction data. It basically accepts the run, and the BU as
    an argument and then determines the period in which tansactions were present. It then creates a sample based
    on this determined period.'''
    
    #Determine the periods in which transacitons are present for the business unit, and create column names
    if BU == cunstructor_dict['aggregate_sheet_name']:
        period_start_date=min(x.TransactionDate for x in TransactionData.objects.filter(run=run))
        period_end_date=max(x.TransactionDate for x in TransactionData.objects.filter(run=run))
    else:
        period_start_date= run.RunPeriodStart #   min(x.TransactionDate for x in TransactionData.objects.filter(run=run, BusinessUnit = BU))
        period_end_date= run.RunPeriodEnd #  max(x.TransactionDate for x in TransactionData.objects.filter(run=run, BusinessUnit = BU))
    date_extent=[period_start_date+timedelta(i) for i in range((period_end_date-period_start_date).days+1)]
    date_set=set(datetime.strftime(i,'%b-%Y') for i in date_extent)
    period_column_names = sorted(date_set, key=lambda day: datetime.strptime(day, "%b-%Y"))
    
    #Create a preamble
    preamble_rows=[['Business Unit', 'Please duplicate this sheet for each business unit'],['Period Start Date',period_start_date ],['Period End Date',period_end_date ],[]]

    #Create the list of column names
    leading_column_names=cunstructor_dict['leading_column_names']
    trailing_column_names=[cunstructor_dict['total_column_name']]
    pnl_column_names=leading_column_names+period_column_names+trailing_column_names    
    
    #Create sample body
    sample_body = [list(i) for i in zip(cunstructor_dict['sample_line_descriptions'],cunstructor_dict['sample_line_codes'])]
    
    #Append preamble to worksheet
    for row in preamble_rows:
        ws.append(row)

    #Append header
    ws.append(pnl_column_names)

    #Append body, marking the aggregate rows as bold
    for i, rowvals in [i for i in zip(cunstructor_dict['sample_line_aggregate_status'],sample_body)]:
        row=[]
        for v in rowvals:
            newcell = cell.Cell(ws,column='A', row=1,value=v)
            row.append(newcell)
        if i==1: makeBold(row)
        ws.append(row)

    #Fix column width for readability
    fixWidth(ws)

    #Apply cost and revenue coloring
    applyPnLStyling(ws,pnl_column_names,cunstructor_dict['revenueMarkers'])
    
    return(ws)
    
    
def generateSamplePnL(run, cunstructor_dict = PnL_settings_dict):
    wb=Workbook()
    wb.remove_sheet(wb.active)
    #BUs = set(x.BusinessUnit for x in TransactionData.objects.filter(run=run))
    #BUs.add(cunstructor_dict['aggregate_sheet_name'])
    BUs = [cunstructor_dict['aggregate_sheet_name']] #well, we decided to create a sample just for the whole thing in the end.
    for BU in BUs:
        ws=wb.create_sheet(BU)
        generateSamplePnLBUSheet(run, ws, BU, cunstructor_dict)
    return(ContentFile(save_virtual_workbook(wb)))
    
    
    
def updateRunPeriod(run):
    run.RunPeriodStart = min(x.TransactionDate for x in TransactionData.objects.filter(run=run))
    run.RunPeriodEnd = max(x.TransactionDate for x in TransactionData.objects.filter(run=run))
    run.save()
    

def returnPnLPeriodColumnNames(run):
    date_extent=[run.RunPeriodStart+timedelta(i) for i in range((run.RunPeriodEnd-run.RunPeriodStart).days+1)]
    date_set=set(datetime.strftime(i,'%b-%Y') for i in date_extent)
    return sorted(date_set, key=lambda day: datetime.strptime(day, "%b-%Y"))
    
    
    
    
    
def processPnLSheet_to_pd(ws, run, PnL_settings_dict=PnL_settings_dict):
    '''This function processes a PnL file sheet. It takes an openpyxl worksheet (ws), and then returnds a pandas dataframe.
    The arguments are:
    1. ws: Name of the worksheet to process
    2. PnL_settings_dict: The dictionary in prof/settings'''
    desiredColumns = PnL_settings_dict['leading_column_names'] + returnPnLPeriodColumnNames(run) + [PnL_settings_dict['total_column_name']]
    revenueSwitch=True
    selectedIndex=None
    for row in ws.iter_rows():
        if selectedIndex:
            if isRowValid(row,selectedIndex):
                newrow=[row[i].value for i in selectedIndex]
                newrow.append(isRowAggregate(row,selectedIndex))
                newrow.append(revenueSwitch)
                if isRevAggregate(row,selectedIndex): revenueSwitch = False
                df=df.append(pd.DataFrame([newrow], columns=outputColumnNames))
        elif ifHeaderReturnIndex(row,desiredColumns):
            selectedIndex=ifHeaderReturnIndex(row, desiredColumns)
            outputColumnNames=[row[i].value for i in selectedIndex]
            outputColumnNames.append(PnL_settings_dict['revenue_or_cost_column_name'])
            outputColumnNames.append(PnL_settings_dict['aggregate_line_flag_column_name'])
            df=pd.DataFrame(columns=outputColumnNames)
    return(df)

@transaction.atomic
def processPnLFile(fileObject, run, PnL_settings_dict=PnL_settings_dict, foreignKeyFileObject=None):
    '''This function processes a PnL file sheet. It takes an PnL file object, and feeds the data to the Django model
    named PnLData.
    The arguments are:
    1. fileObject: The PnL file object (an object of FileSpace model)
    2. run: The associated run, an object of the RunSpace model
    3. PnL_settings_dict: The dictionary in prof/settings
    4. foreignKeyFileObject: The file object which should be marked as the foreign key if not the same as fileObject'''
    if not(foreignKeyFileObject): foreignKeyFileObject = fileObject
    wb=load_workbook(fileObject.File, read_only=True, data_only=True)
    periodColumns = returnPnLPeriodColumnNames(run)
    desiredColumns = PnL_settings_dict['leading_column_names'] + returnPnLPeriodColumnNames(run) + [PnL_settings_dict['total_column_name']]
    columnsForModel = PnL_settings_dict['leading_column_names'] +[PnL_settings_dict['revenue_or_cost_column_name']] + [PnL_settings_dict['aggregate_line_flag_column_name']]
    fullPnL=pd.concat([processPnLSheet_to_pd(ws, run, PnL_settings_dict).set_index(PnL_settings_dict['leading_column_names']) for ws in wb.worksheets], keys=wb.get_sheet_names(),names=[PnL_settings_dict['BU_column_name']]).reset_index().replace(np.nan,0)
    fullPnL = pd.melt(fullPnL,id_vars=columnsForModel, value_vars=       
        periodColumns,
            var_name = PnL_settings_dict['period_column_name'], value_name=PnL_settings_dict['amount_column_name'])
    fullPnL.columns = [PnL_settings_dict['column_names_to_model_dict'][i] for i in fullPnL.columns]
    list_of_dicts = [row.to_dict() for index,row in fullPnL.iterrows()]
    for row in list_of_dicts:
        row['run'] = run
        row['SourceFile'] = foreignKeyFileObject
        PnLData(**row).save()
        
        
